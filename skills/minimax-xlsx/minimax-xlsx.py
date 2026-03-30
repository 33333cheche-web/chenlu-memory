#!/usr/bin/env python3
"""
MiniMax-xlsx Skill - Excel 创建和编辑
支持：创建、读取、编辑、公式验证
"""

import argparse
import json
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_excel(data, output_path):
    """创建Excel文件"""
    wb = Workbook()
    ws = wb.active
    
    # 写入数据
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 表头样式
            if row_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1A4D44", end_color="1A4D44", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    return output_path

def main():
    parser = argparse.ArgumentParser(description='MiniMax-xlsx: Excel创建和编辑')
    parser.add_argument('--create', '-c', action='store_true', help='创建新Excel')
    parser.add_argument('--data', '-d', type=str, help='JSON格式数据')
    parser.add_argument('--output', '-o', type=str, required=True, help='输出路径')
    
    args = parser.parse_args()
    
    if args.create:
        data = json.loads(args.data) if args.data else [["示例", "数据"]]
        output_path = create_excel(data, args.output)
        print(f"✅ Excel已创建: {output_path}")

if __name__ == "__main__":
    main()
